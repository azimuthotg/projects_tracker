import io

import openpyxl
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle
from reportlab.lib.units import cm
from reportlab.platypus import (
    HRFlowable,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
)
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from django.contrib.auth.decorators import login_required
from django.db.models import FloatField, Q, Sum
from django.db.models.functions import Cast
from django.http import HttpResponse
from django.shortcuts import get_object_or_404, render
from django.utils import timezone

from apps.accounts.models import Department
from .pdf_utils import (
    build_document,
    fmt_currency,
    header_block,
    make_styles,
    pct_color,
    summary_card_row,
    table_header_style,
    thaidate as pdf_thaidate,
    C_ALT_ROW, C_AMBER, C_BLUE, C_BORDER, C_EMERALD, C_HEADER_BG,
    C_RED, C_SUBHEAD_BG, C_SUBHEAD_FG, C_TOTAL_BG,
)
from apps.budget.models import Expense
from apps.projects.models import Activity, FiscalYear, Project
from apps.projects.utils import get_projects_for_user


# ─── Helpers ────────────────────────────────────────────────────────────────

def _apply_header_style(cell, bg="1e3a5f", fg="FFFFFF", bold=True, size=10):
    cell.font = Font(bold=bold, color=fg, size=size)
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def _apply_subheader_style(cell, bg="dbeafe"):
    cell.font = Font(bold=True, size=9, color="1e3a5f")
    cell.fill = PatternFill("solid", fgColor=bg)
    cell.alignment = Alignment(horizontal="center", vertical="center")


def _thin_border():
    s = Side(style="thin", color="cccccc")
    return Border(left=s, right=s, top=s, bottom=s)


def _currency(val):
    try:
        return float(val or 0)
    except Exception:
        return 0.0


def _thai_month(m):
    names = ['', 'ม.ค.', 'ก.พ.', 'มี.ค.', 'เม.ย.', 'พ.ค.', 'มิ.ย.',
             'ก.ค.', 'ส.ค.', 'ก.ย.', 'ต.ค.', 'พ.ย.', 'ธ.ค.']
    return names[m] if 1 <= m <= 12 else ''


def _thaidate(d):
    if not d:
        return ''
    return f"{d.day} {_thai_month(d.month)} {d.year + 543}"


# ─── 1. Budget Overview ──────────────────────────────────────────────────────

@login_required
def budget_report(request):
    fiscal_years = FiscalYear.objects.all().order_by('-year')
    departments = Department.objects.all().order_by('name')

    fy_id = request.GET.get('fiscal_year')
    dept_id = request.GET.get('department')
    status = request.GET.get('status')

    projects = get_projects_for_user(request.user)

    if fy_id:
        projects = projects.filter(fiscal_year_id=fy_id)
    else:
        active_fy = fiscal_years.filter(is_active=True).first()
        if active_fy:
            fy_id = str(active_fy.pk)
            projects = projects.filter(fiscal_year=active_fy)

    if dept_id:
        projects = projects.filter(department_id=dept_id)
    if status:
        projects = projects.filter(status=status)

    projects = projects.prefetch_related(
        'budget_sources', 'activities', 'department'
    ).annotate(
        code_num=Cast('project_code', FloatField())
    ).order_by('code_num', 'project_code')

    # Summary totals
    total_budget = sum(_currency(p.total_budget) for p in projects)
    total_spent = sum(_currency(p.total_spent) for p in projects)
    total_remaining = total_budget - total_spent
    total_pct = (total_spent / total_budget * 100) if total_budget > 0 else 0

    fiscal_year = fiscal_years.filter(pk=fy_id).first() if fy_id else None

    return render(request, 'reports/budget_report.html', {
        'projects': projects,
        'fiscal_years': fiscal_years,
        'departments': departments,
        'status_choices': Project.STATUS_CHOICES,
        'current_fy': fy_id or '',
        'current_dept': dept_id or '',
        'current_status': status or '',
        'fiscal_year': fiscal_year,
        'total_budget': total_budget,
        'total_spent': total_spent,
        'total_remaining': total_remaining,
        'total_pct': total_pct,
        'today': timezone.now().date(),
    })


@login_required
def budget_report_print(request):
    fiscal_years = FiscalYear.objects.all().order_by('-year')
    departments = Department.objects.all().order_by('name')

    fy_id = request.GET.get('fiscal_year')
    dept_id = request.GET.get('department')
    status = request.GET.get('status')

    projects = get_projects_for_user(request.user)

    if fy_id:
        projects = projects.filter(fiscal_year_id=fy_id)
    else:
        active_fy = fiscal_years.filter(is_active=True).first()
        if active_fy:
            fy_id = str(active_fy.pk)
            projects = projects.filter(fiscal_year=active_fy)

    if dept_id:
        projects = projects.filter(department_id=dept_id)
    if status:
        projects = projects.filter(status=status)

    projects = projects.prefetch_related(
        'budget_sources', 'activities', 'department'
    ).annotate(
        code_num=Cast('project_code', FloatField())
    ).order_by('code_num', 'project_code')

    total_budget = sum(_currency(p.total_budget) for p in projects)
    total_spent = sum(_currency(p.total_spent) for p in projects)
    total_remaining = total_budget - total_spent
    total_pct = (total_spent / total_budget * 100) if total_budget > 0 else 0

    fiscal_year = fiscal_years.filter(pk=fy_id).first() if fy_id else None
    dept_label = departments.filter(pk=dept_id).first().name if dept_id else ''
    status_label = dict(Project.STATUS_CHOICES).get(status, '') if status else ''

    return render(request, 'reports/budget_report_print.html', {
        'projects': projects,
        'fiscal_year': fiscal_year,
        'dept_label': dept_label,
        'status_label': status_label,
        'total_budget': total_budget,
        'total_spent': total_spent,
        'total_remaining': total_remaining,
        'total_pct': total_pct,
        'today': timezone.now().date(),
    })


@login_required
def budget_report_excel(request):
    fiscal_years = FiscalYear.objects.all().order_by('-year')
    fy_id = request.GET.get('fiscal_year')
    dept_id = request.GET.get('department')
    status = request.GET.get('status')

    projects = get_projects_for_user(request.user)
    if fy_id:
        projects = projects.filter(fiscal_year_id=fy_id)
    else:
        active_fy = fiscal_years.filter(is_active=True).first()
        if active_fy:
            projects = projects.filter(fiscal_year=active_fy)
    if dept_id:
        projects = projects.filter(department_id=dept_id)
    if status:
        projects = projects.filter(status=status)

    projects = projects.prefetch_related('budget_sources', 'activities', 'department').annotate(
        code_num=Cast('project_code', FloatField())
    ).order_by('code_num', 'project_code')

    fiscal_year = fiscal_years.filter(pk=fy_id).first() if fy_id else fiscal_years.filter(is_active=True).first()
    fy_label = f"ปีงบประมาณ {fiscal_year.year}" if fiscal_year else "ทุกปีงบประมาณ"

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "รายงานงบประมาณ"

    border = _thin_border()
    num_fmt = '#,##0.00'
    pct_fmt = '0.0"%"'

    # Title
    ws.merge_cells('A1:J1')
    t = ws['A1']
    t.value = f"รายงานภาพรวมงบประมาณ — {fy_label}"
    t.font = Font(bold=True, size=14, color="1e3a5f")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells('A2:J2')
    ws['A2'].value = f"ออกรายงาน: {_thaidate(timezone.now().date())}"
    ws['A2'].font = Font(size=9, color="888888")
    ws['A2'].alignment = Alignment(horizontal="center")

    # Headers
    headers = [
        'รหัส', 'ชื่อโครงการ', 'แผนก', 'สถานะ',
        'เงินแผ่นดิน', 'เงินสะสม', 'เงินรายได้',
        'งบรวม', 'ใช้ไป', 'คงเหลือ', '% ใช้',
    ]
    col_widths = [10, 40, 20, 14, 16, 16, 16, 16, 16, 16, 10]

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=ci, value=h)
        _apply_header_style(cell)
        cell.border = border
        ws.column_dimensions[get_column_letter(ci)].width = col_widths[ci - 1]
    ws.row_dimensions[4].height = 22

    # Data
    row_num = 5
    total_row = {k: 0.0 for k in ['gov', 'acc', 'rev', 'total', 'spent', 'remaining']}

    for p in projects:
        bs = p.budget_by_source
        gov = _currency(bs.get('government', 0))
        acc = _currency(bs.get('accumulated', 0))
        rev = _currency(bs.get('revenue', 0))
        total = _currency(p.total_budget)
        spent = _currency(p.total_spent)
        remaining = total - spent
        pct = (spent / total * 100) if total > 0 else 0

        total_row['gov'] += gov
        total_row['acc'] += acc
        total_row['rev'] += rev
        total_row['total'] += total
        total_row['spent'] += spent
        total_row['remaining'] += remaining

        row_data = [
            p.project_code,
            p.name,
            p.department.name if p.department else '',
            p.get_status_display(),
            gov, acc, rev, total, spent, remaining, pct,
        ]
        bg = "f8faff" if row_num % 2 == 0 else "ffffff"
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=ci, value=val)
            cell.border = border
            cell.font = Font(size=9)
            cell.fill = PatternFill("solid", fgColor=bg)
            if ci >= 5:
                cell.number_format = pct_fmt if ci == 11 else num_fmt
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(vertical="center", wrap_text=(ci == 2))
        row_num += 1

    # Total row
    total_pct = (total_row['spent'] / total_row['total'] * 100) if total_row['total'] > 0 else 0
    total_data = ['', 'รวมทั้งหมด', '', '',
                  total_row['gov'], total_row['acc'], total_row['rev'],
                  total_row['total'], total_row['spent'], total_row['remaining'], total_pct]
    for ci, val in enumerate(total_data, 1):
        cell = ws.cell(row=row_num, column=ci, value=val)
        cell.font = Font(bold=True, size=9, color="1e3a5f")
        cell.fill = PatternFill("solid", fgColor="dbeafe")
        cell.border = border
        if ci >= 5:
            cell.number_format = pct_fmt if ci == 11 else num_fmt
            cell.alignment = Alignment(horizontal="right")

    ws.freeze_panes = 'A5'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="budget_report.xlsx"'
    return response


# ─── 2. Expense Report ───────────────────────────────────────────────────────

@login_required
def expense_report(request):
    fiscal_years = FiscalYear.objects.all().order_by('-year')
    departments = Department.objects.all().order_by('name')

    fy_id = request.GET.get('fiscal_year')
    dept_id = request.GET.get('department')
    project_id = request.GET.get('project')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    projects_qs = get_projects_for_user(request.user)

    if fy_id:
        projects_qs = projects_qs.filter(fiscal_year_id=fy_id)
    else:
        active_fy = fiscal_years.filter(is_active=True).first()
        if active_fy:
            fy_id = str(active_fy.pk)
            projects_qs = projects_qs.filter(fiscal_year=active_fy)

    if dept_id:
        projects_qs = projects_qs.filter(department_id=dept_id)

    expenses = Expense.objects.filter(
        activity__project__in=projects_qs,
        status='approved',
    ).select_related(
        'activity', 'activity__project', 'activity__project__department',
        'created_by', 'approved_by',
    ).order_by('activity__project__project_code', 'activity__activity_number', 'expense_date')

    if project_id:
        expenses = expenses.filter(activity__project_id=project_id)
    if date_from:
        expenses = expenses.filter(expense_date__gte=date_from)
    if date_to:
        expenses = expenses.filter(expense_date__lte=date_to)

    total_amount = expenses.aggregate(total=Sum('amount'))['total'] or 0

    # Projects for filter dropdown
    available_projects = projects_qs.order_by('project_code')
    fiscal_year = fiscal_years.filter(pk=fy_id).first() if fy_id else None

    return render(request, 'reports/expense_report.html', {
        'expenses': expenses,
        'fiscal_years': fiscal_years,
        'departments': departments,
        'available_projects': available_projects,
        'current_fy': fy_id or '',
        'current_dept': dept_id or '',
        'current_project': project_id or '',
        'current_date_from': date_from or '',
        'current_date_to': date_to or '',
        'fiscal_year': fiscal_year,
        'total_amount': total_amount,
    })


@login_required
def expense_report_excel(request):
    fiscal_years = FiscalYear.objects.all().order_by('-year')
    fy_id = request.GET.get('fiscal_year')
    dept_id = request.GET.get('department')
    project_id = request.GET.get('project')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')

    projects_qs = get_projects_for_user(request.user)
    if fy_id:
        projects_qs = projects_qs.filter(fiscal_year_id=fy_id)
    else:
        active_fy = fiscal_years.filter(is_active=True).first()
        if active_fy:
            projects_qs = projects_qs.filter(fiscal_year=active_fy)
    if dept_id:
        projects_qs = projects_qs.filter(department_id=dept_id)

    expenses = Expense.objects.filter(
        activity__project__in=projects_qs, status='approved',
    ).select_related(
        'activity', 'activity__project', 'activity__project__department', 'created_by',
    ).order_by('activity__project__project_code', 'activity__activity_number', 'expense_date')

    if project_id:
        expenses = expenses.filter(activity__project_id=project_id)
    if date_from:
        expenses = expenses.filter(expense_date__gte=date_from)
    if date_to:
        expenses = expenses.filter(expense_date__lte=date_to)

    fiscal_year = fiscal_years.filter(pk=fy_id).first() if fy_id else fiscal_years.filter(is_active=True).first()
    fy_label = f"ปีงบประมาณ {fiscal_year.year}" if fiscal_year else ""

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "รายงานเบิกจ่าย"
    border = _thin_border()
    num_fmt = '#,##0.00'

    ws.merge_cells('A1:I1')
    t = ws['A1']
    t.value = f"รายงานการเบิกจ่าย — {fy_label}"
    t.font = Font(bold=True, size=14, color="1e3a5f")
    t.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 28

    ws.merge_cells('A2:I2')
    ws['A2'].value = f"ออกรายงาน: {_thaidate(timezone.now().date())}"
    ws['A2'].font = Font(size=9, color="888888")
    ws['A2'].alignment = Alignment(horizontal="center")

    headers = ['รหัสโครงการ', 'ชื่อโครงการ', 'กิจกรรม', 'รายการ',
               'เลขที่ใบเสร็จ', 'วันที่', 'แหล่งเงิน', 'จำนวนเงิน', 'อนุมัติโดย']
    col_widths = [14, 35, 30, 35, 16, 14, 14, 16, 20]

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=4, column=ci, value=h)
        _apply_header_style(cell)
        cell.border = border
        ws.column_dimensions[get_column_letter(ci)].width = col_widths[ci - 1]
    ws.row_dimensions[4].height = 22

    SOURCE_LABELS = {'government': 'เงินแผ่นดิน', 'accumulated': 'เงินสะสม', 'revenue': 'เงินรายได้'}
    row_num = 5
    grand_total = 0.0

    for i, exp in enumerate(expenses):
        amt = _currency(exp.amount)
        grand_total += amt
        row_data = [
            exp.activity.project.project_code,
            exp.activity.project.name,
            f"{exp.activity.activity_number}. {exp.activity.name}",
            exp.description,
            exp.receipt_number or '',
            _thaidate(exp.expense_date),
            SOURCE_LABELS.get(exp.budget_source, exp.budget_source or ''),
            amt,
            exp.approved_by.get_full_name() if exp.approved_by else '',
        ]
        bg = "f8faff" if i % 2 == 0 else "ffffff"
        for ci, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=ci, value=val)
            cell.border = border
            cell.font = Font(size=9)
            cell.fill = PatternFill("solid", fgColor=bg)
            if ci == 8:
                cell.number_format = num_fmt
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.alignment = Alignment(vertical="center", wrap_text=(ci in [2, 3, 4]))
        row_num += 1

    # Grand total
    ws.merge_cells(f'A{row_num}:G{row_num}')
    cell = ws.cell(row=row_num, column=1, value='รวมทั้งหมด')
    cell.font = Font(bold=True, size=9, color="1e3a5f")
    cell.fill = PatternFill("solid", fgColor="dbeafe")
    cell.alignment = Alignment(horizontal="right")
    cell.border = border

    total_cell = ws.cell(row=row_num, column=8, value=grand_total)
    total_cell.font = Font(bold=True, size=9, color="1e3a5f")
    total_cell.fill = PatternFill("solid", fgColor="dbeafe")
    total_cell.number_format = num_fmt
    total_cell.alignment = Alignment(horizontal="right")
    total_cell.border = border

    ws.freeze_panes = 'A5'

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    response = HttpResponse(
        buf.read(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="expense_report.xlsx"'
    return response


# ─── 3. Project Detail Report (Print/PDF) ────────────────────────────────────

@login_required
def project_report(request, pk):
    projects = get_projects_for_user(request.user)
    project = get_object_or_404(projects, pk=pk)

    activities = project.activities.prefetch_related(
        'responsible_persons'
    ).order_by('activity_number')

    expenses = Expense.objects.filter(
        activity__project=project, status='approved',
    ).select_related('activity', 'created_by').order_by(
        'activity__activity_number', 'expense_date'
    )

    total_spent = expenses.aggregate(total=Sum('amount'))['total'] or 0

    return render(request, 'reports/project_report.html', {
        'project': project,
        'activities': activities,
        'expenses': expenses,
        'total_spent': total_spent,
        'now': timezone.now(),
    })


# ─── 4. Budget Overview PDF ──────────────────────────────────────────────────

@login_required
def budget_report_pdf(request):
    fiscal_years = FiscalYear.objects.all().order_by('-year')
    fy_id = request.GET.get('fiscal_year')
    dept_id = request.GET.get('department')
    status = request.GET.get('status')

    projects = get_projects_for_user(request.user)
    if fy_id:
        projects = projects.filter(fiscal_year_id=fy_id)
    else:
        active_fy = fiscal_years.filter(is_active=True).first()
        if active_fy:
            fy_id = str(active_fy.pk)
            projects = projects.filter(fiscal_year=active_fy)
    if dept_id:
        projects = projects.filter(department_id=dept_id)
    if status:
        projects = projects.filter(status=status)

    projects = projects.prefetch_related(
        'budget_sources', 'activities', 'department'
    ).annotate(
        code_num=Cast('project_code', FloatField())
    ).order_by('code_num', 'project_code')

    fiscal_year = fiscal_years.filter(pk=fy_id).first() if fy_id else None
    dept_obj = Department.objects.filter(pk=dept_id).first() if dept_id else None

    project_list = list(projects)
    total_budget = sum(_currency(p.total_budget) for p in project_list)
    total_spent = sum(_currency(p.total_spent) for p in project_list)
    total_remaining = total_budget - total_spent
    total_pct = (total_spent / total_budget * 100) if total_budget > 0 else 0

    # Build subtitle from filters
    parts = []
    if fiscal_year:
        parts.append(f"ปีงบประมาณ {fiscal_year.year}")
    if dept_obj:
        parts.append(dept_obj.name)
    if status:
        parts.append(f"สถานะ: {dict(Project.STATUS_CHOICES).get(status, status)}")
    parts.append(f"ออกรายงาน: {pdf_thaidate(timezone.now().date())}")
    subtitle = "  |  ".join(parts)

    styles = make_styles()
    buf = io.BytesIO()
    doc = build_document(buf, "รายงานภาพรวมงบประมาณ", page_size=landscape(A4))

    W = landscape(A4)[0] - 3 * cm  # usable width (landscape)

    story = []

    # Header block
    story.append(header_block(styles, "รายงานภาพรวมงบประมาณ", subtitle, page_width=W))
    story.append(Spacer(1, 6))

    # Summary cards
    cards = [
        ("โครงการทั้งหมด", str(len(project_list)), "โครงการ"),
        ("งบประมาณรวม", fmt_currency(total_budget), "บาท"),
        ("ใช้ไปแล้ว", fmt_currency(total_spent), f"บาท ({total_pct:.1f}%)"),
        ("คงเหลือ", fmt_currency(total_remaining), "บาท"),
    ]
    story.append(summary_card_row(styles, cards, W))
    story.append(Spacer(1, 8))

    # Table
    STATUS_TH = {
        'draft': 'ร่าง', 'not_started': 'ยังไม่เริ่ม',
        'active': 'ดำเนินการ', 'completed': 'เสร็จสิ้น', 'cancelled': 'ยกเลิก',
    }

    col_widths = [1.2*cm, 9.5*cm, 2.0*cm, 2.2*cm, 2.2*cm, 2.2*cm, 2.5*cm, 2.5*cm, 2.5*cm, 1.9*cm]
    headers = [
        Paragraph("รหัส", styles['th']),
        Paragraph("ชื่อโครงการ", styles['th']),
        Paragraph("สถานะ", styles['th']),
        Paragraph("เงิน\nแผ่นดิน", styles['th']),
        Paragraph("เงิน\nสะสม", styles['th']),
        Paragraph("เงิน\nรายได้", styles['th']),
        Paragraph("งบรวม", styles['th']),
        Paragraph("ใช้ไป", styles['th']),
        Paragraph("คงเหลือ", styles['th']),
        Paragraph("%", styles['th']),
    ]
    data = [headers]

    project_list = list(projects)
    for i, p in enumerate(project_list):
        bs = p.budget_by_source
        gov  = _currency(bs.get('government', 0))
        acc  = _currency(bs.get('accumulated', 0))
        rev  = _currency(bs.get('revenue', 0))
        total = _currency(p.total_budget)
        spent = _currency(p.total_spent)
        remaining = total - spent
        pct = (spent / total * 100) if total > 0 else 0.0

        pc = pct_color(pct)
        name_cell = [
            Paragraph(p.name, styles['td']),
            Paragraph(p.department.name if p.department else '', styles['small']),
        ]
        row = [
            Paragraph(p.project_code or '', styles['td_c']),
            name_cell,
            Paragraph(STATUS_TH.get(p.status, p.status), styles['td_c']),
            Paragraph(fmt_currency(gov) if gov else '—', styles['td_r']),
            Paragraph(fmt_currency(acc) if acc else '—', styles['td_r']),
            Paragraph(fmt_currency(rev) if rev else '—', styles['td_r']),
            Paragraph(fmt_currency(total), styles['td_bold_r']),
            Paragraph(fmt_currency(spent), ParagraphStyle('sp', fontName='THSarabunNew',
                      fontSize=9, alignment=2, textColor=pc)),
            Paragraph(fmt_currency(remaining), ParagraphStyle('rem', fontName='THSarabunNew',
                      fontSize=9, alignment=2, textColor=C_EMERALD)),
            Paragraph(f"{pct:.1f}%", ParagraphStyle('pct', fontName='THSarabunNew-Bold',
                      fontSize=9, alignment=2, textColor=pc)),
        ]
        data.append(row)

    # Total row
    total_pct2 = (total_spent / total_budget * 100) if total_budget > 0 else 0
    data.append([
        Paragraph('', styles['td']),
        Paragraph('รวมทั้งหมด', styles['td_bold']),
        Paragraph('', styles['td']),
        Paragraph('', styles['td']),
        Paragraph('', styles['td']),
        Paragraph('', styles['td']),
        Paragraph(fmt_currency(total_budget), styles['td_bold_r']),
        Paragraph(fmt_currency(total_spent), styles['td_bold_r']),
        Paragraph(fmt_currency(total_remaining), styles['td_bold_r']),
        Paragraph(f"{total_pct2:.1f}%", styles['td_bold_r']),
    ])

    tbl = Table(data, colWidths=col_widths, repeatRows=1)

    # Build row alternating background + total row style
    n = len(data)
    extra_cmds = []
    for row_idx in range(1, n - 1):
        if row_idx % 2 == 0:
            extra_cmds.append(('BACKGROUND', (0, row_idx), (-1, row_idx), C_ALT_ROW))
    # Total row
    extra_cmds += [
        ('BACKGROUND', (0, n - 1), (-1, n - 1), C_TOTAL_BG),
        ('FONTNAME',   (0, n - 1), (-1, n - 1), 'THSarabunNew-Bold'),
        ('LINEABOVE',  (0, n - 1), (-1, n - 1), 1, C_HEADER_BG),
    ]
    tbl.setStyle(table_header_style(len(col_widths), extra=extra_cmds))
    story.append(tbl)

    # Footer
    story.append(Spacer(1, 10))
    story.append(HRFlowable(width=W, thickness=0.5, color=C_BORDER))
    story.append(Spacer(1, 4))
    story.append(Paragraph(
        f"ระบบติดตามแผนงาน โครงการและงบประมาณ — สำนักวิทยบริการ มหาวิทยาลัยนครพนม  |  {pdf_thaidate(timezone.now().date())}",
        styles['footer']
    ))

    doc.build(story)
    buf.seek(0)

    fy_label = f"_{fiscal_year.year}" if fiscal_year else ""
    response = HttpResponse(buf.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="budget_report{fy_label}.pdf"'
    return response


# ─── 5. Project Detail PDF ───────────────────────────────────────────────────

@login_required
def project_report_pdf(request, pk):
    projects_qs = get_projects_for_user(request.user)
    project = get_object_or_404(projects_qs, pk=pk)

    activities = project.activities.prefetch_related(
        'responsible_persons'
    ).order_by('activity_number')

    expenses = Expense.objects.filter(
        activity__project=project, status='approved',
    ).select_related('activity', 'created_by', 'approved_by').order_by(
        'activity__activity_number', 'expense_date'
    )

    total_spent = expenses.aggregate(total=Sum('amount'))['total'] or 0
    total_budget = _currency(project.total_budget)
    total_remaining = total_budget - _currency(total_spent)
    total_pct = (_currency(total_spent) / total_budget * 100) if total_budget > 0 else 0

    styles = make_styles()
    buf = io.BytesIO()
    doc = build_document(buf, f"รายงานโครงการ — {project.name}")

    W = A4[0] - 3 * cm

    STATUS_TH = {
        'draft': 'ร่าง', 'not_started': 'ยังไม่เริ่ม',
        'active': 'ดำเนินการ', 'completed': 'เสร็จสิ้น', 'cancelled': 'ยกเลิก',
    }
    ACTIVITY_STATUS_TH = {
        'pending': 'รอดำเนินการ', 'in_progress': 'กำลังดำเนินการ',
        'completed': 'เสร็จสิ้น', 'cancelled': 'ยกเลิก',
    }
    SOURCE_LABELS = {'government': 'เงินแผ่นดิน', 'accumulated': 'เงินสะสม', 'revenue': 'เงินรายได้'}

    bs = project.budget_by_source
    subtitle_parts = [
        f"รหัส: {project.project_code}" if project.project_code else '',
        STATUS_TH.get(project.status, project.status),
        f"ออกรายงาน: {pdf_thaidate(timezone.now().date())}",
    ]
    subtitle = "  |  ".join(p for p in subtitle_parts if p)

    story = []
    story.append(header_block(styles, project.name, subtitle))
    story.append(Spacer(1, 6))

    # Budget summary cards
    cards = [
        ("งบประมาณรวม", fmt_currency(total_budget), "บาท"),
        ("ใช้ไปแล้ว", fmt_currency(total_spent), f"บาท ({total_pct:.1f}%)"),
        ("คงเหลือ", fmt_currency(total_remaining), "บาท"),
        ("ช่วงเวลา",
         f"{pdf_thaidate(project.start_date)} – {pdf_thaidate(project.end_date)}" if project.start_date else '—',
         ''),
    ]
    story.append(summary_card_row(styles, cards, W))
    story.append(Spacer(1, 12))

    # ── Activities Table ────────────────────────────────────────────────────

    story.append(Paragraph("กิจกรรมโครงการ", ParagraphStyle(
        'sec', fontName='THSarabunNew-Bold', fontSize=12, leading=16,
        textColor=C_HEADER_BG)))
    story.append(Spacer(1, 4))

    act_widths = [0.8*cm, 5.0*cm, 1.8*cm, 2.0*cm, 2.0*cm, 2.2*cm, 2.2*cm, 2.2*cm, 2.0*cm]
    act_headers = [
        Paragraph("ที่", styles['th']),
        Paragraph("ชื่อกิจกรรม", styles['th']),
        Paragraph("สถานะ", styles['th']),
        Paragraph("เริ่มต้น", styles['th']),
        Paragraph("สิ้นสุด", styles['th']),
        Paragraph("งบที่ได้รับ", styles['th']),
        Paragraph("ใช้ไป", styles['th']),
        Paragraph("คงเหลือ", styles['th']),
        Paragraph("%", styles['th']),
    ]
    act_data = [act_headers]

    for i, act in enumerate(activities):
        a_spent = _currency(act.total_spent)
        a_budget = _currency(act.allocated_budget)
        a_remaining = a_budget - a_spent
        a_pct = (a_spent / a_budget * 100) if a_budget > 0 else 0.0
        pc = pct_color(a_pct)
        act_data.append([
            Paragraph(str(act.activity_number), styles['td_c']),
            Paragraph(act.name, styles['td']),
            Paragraph(ACTIVITY_STATUS_TH.get(act.status, act.status), styles['td_c']),
            Paragraph(pdf_thaidate(act.start_date), styles['td_c']),
            Paragraph(pdf_thaidate(act.end_date), styles['td_c']),
            Paragraph(fmt_currency(a_budget), styles['td_r']),
            Paragraph(fmt_currency(a_spent), ParagraphStyle('as', fontName='THSarabunNew',
                      fontSize=9, alignment=2, textColor=pc)),
            Paragraph(fmt_currency(a_remaining), ParagraphStyle('ar', fontName='THSarabunNew',
                      fontSize=9, alignment=2, textColor=C_EMERALD)),
            Paragraph(f"{a_pct:.1f}%", ParagraphStyle('ap', fontName='THSarabunNew-Bold',
                      fontSize=9, alignment=2, textColor=pc)),
        ])

    act_n = len(act_data)
    act_extra = []
    for ri in range(1, act_n):
        if ri % 2 == 0:
            act_extra.append(('BACKGROUND', (0, ri), (-1, ri), C_ALT_ROW))

    act_tbl = Table(act_data, colWidths=act_widths, repeatRows=1)
    act_tbl.setStyle(table_header_style(len(act_widths), extra=act_extra))
    story.append(act_tbl)
    story.append(Spacer(1, 12))

    # ── Expenses Table ──────────────────────────────────────────────────────

    story.append(Paragraph("รายการเบิกจ่าย (อนุมัติแล้ว)", ParagraphStyle(
        'sec2', fontName='THSarabunNew-Bold', fontSize=12, leading=16,
        textColor=C_HEADER_BG)))
    story.append(Spacer(1, 4))

    exp_widths = [2.8*cm, 4.8*cm, 3.2*cm, 2.2*cm, 2.2*cm, 2.4*cm, 2.6*cm]
    exp_headers = [
        Paragraph("กิจกรรม", styles['th']),
        Paragraph("รายการ", styles['th']),
        Paragraph("เลขที่ใบเสร็จ", styles['th']),
        Paragraph("วันที่", styles['th']),
        Paragraph("แหล่งเงิน", styles['th']),
        Paragraph("จำนวนเงิน", styles['th']),
        Paragraph("อนุมัติโดย", styles['th']),
    ]
    exp_data = [exp_headers]

    for i, exp in enumerate(expenses):
        exp_data.append([
            Paragraph(f"{exp.activity.activity_number}. {exp.activity.name}", styles['td']),
            Paragraph(exp.description, styles['td']),
            Paragraph(exp.receipt_number or '—', styles['td_c']),
            Paragraph(pdf_thaidate(exp.expense_date), styles['td_c']),
            Paragraph(SOURCE_LABELS.get(exp.budget_source, exp.budget_source or '—'), styles['td_c']),
            Paragraph(fmt_currency(exp.amount), styles['td_r']),
            Paragraph(exp.approved_by.get_full_name() if exp.approved_by else '—', styles['td']),
        ])

    # Grand total row
    exp_n = len(exp_data)
    exp_data.append([
        Paragraph('', styles['td']),
        Paragraph('รวมทั้งหมด', styles['td_bold']),
        Paragraph('', styles['td']),
        Paragraph('', styles['td']),
        Paragraph('', styles['td']),
        Paragraph(fmt_currency(total_spent), styles['td_bold_r']),
        Paragraph('', styles['td']),
    ])

    exp_extra = []
    for ri in range(1, exp_n):
        if ri % 2 == 0:
            exp_extra.append(('BACKGROUND', (0, ri), (-1, ri), C_ALT_ROW))
    exp_extra += [
        ('BACKGROUND', (0, exp_n), (-1, exp_n), C_TOTAL_BG),
        ('FONTNAME',   (0, exp_n), (-1, exp_n), 'THSarabunNew-Bold'),
        ('LINEABOVE',  (0, exp_n), (-1, exp_n), 1, C_HEADER_BG),
    ]

    exp_tbl = Table(exp_data, colWidths=exp_widths, repeatRows=1)
    exp_tbl.setStyle(table_header_style(len(exp_widths), extra=exp_extra))
    story.append(exp_tbl)

    # Footer
    story.append(Spacer(1, 10))
    story.append(HRFlowable(width=W, thickness=0.5, color=C_BORDER))
    story.append(Spacer(1, 4))
    story.append(Paragraph(
        f"ระบบติดตามแผนงาน โครงการและงบประมาณ — สำนักวิทยบริการ มหาวิทยาลัยนครพนม  |  {pdf_thaidate(timezone.now().date())}",
        styles['footer']
    ))

    doc.build(story)
    buf.seek(0)

    safe_code = (project.project_code or str(pk)).replace('/', '_')
    response = HttpResponse(buf.read(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="project_{safe_code}.pdf"'
    return response
